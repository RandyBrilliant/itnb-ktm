"""Build an .xlsx participant list for a webinar (admin download)."""

from __future__ import annotations

import io

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    "No",
    "Full name",
    "Institutional ID",
    "Email",
    "Department",
    "Status",
    "Registered at",
    "Checked in at",
    "Checked out at",
    "Attended",
    "Certificate issued",
]


def _fmt(dt) -> str:
    if not dt:
        return ""
    return timezone.localtime(dt).strftime("%Y-%m-%d %H:%M")


def build_webinar_participants_workbook(webinar) -> io.BytesIO:
    """Return an in-memory .xlsx of all registrations for the webinar."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="AF0F24")
    ws.append(HEADERS)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 16)

    registrations = (
        webinar.registrations.select_related("user", "certificate")
        .order_by("user__full_name", "registered_at")
    )
    for idx, reg in enumerate(registrations, start=1):
        user = reg.user
        ws.append(
            [
                idx,
                user.full_name or "",
                user.institutional_id or "",
                user.email,
                user.department or "",
                reg.get_status_display(),
                _fmt(reg.registered_at),
                _fmt(reg.checked_in_at),
                _fmt(reg.checked_out_at),
                "Yes" if reg.attended else "No",
                "Yes" if reg.certificate_id else "No",
            ]
        )

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
