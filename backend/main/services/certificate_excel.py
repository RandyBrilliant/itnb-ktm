"""Parse recipient spreadsheets for certificate batches."""

from __future__ import annotations

import re
from dataclasses import dataclass

from openpyxl import load_workbook

NAME_HEADERS = frozenset(
    {
        "name",
        "nama",
        "full name",
        "nama lengkap",
        "participant name",
        "nama peserta",
    }
)
ID_HEADERS = frozenset(
    {
        "id",
        "nim",
        "nip",
        "student id",
        "staff id",
        "nomor",
        "nomor id",
        "no",
        "no.",
        "identity",
    }
)


@dataclass(frozen=True)
class RecipientRow:
    display_name: str
    id_raw: str


def _normalize_header(cell_value) -> str:
    if cell_value is None:
        return ""
    s = str(cell_value).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def parse_recipients_xlsx(file_path: str, *, max_rows: int = 50_000) -> tuple[list[RecipientRow], list[str]]:
    """
    Read the first worksheet and detect Name / ID columns by header row.
    Returns (rows, errors). Duplicate IDs in the sheet are reported as errors (first wins).
    """
    errors: list[str] = []
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header_row_idx = None
        name_col = None
        id_col = None

        for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
            if not row:
                continue
            candidates_name = {}
            candidates_id = {}
            for ci, cell in enumerate(row):
                h = _normalize_header(cell)
                if h in NAME_HEADERS:
                    candidates_name[h] = ci
                if h in ID_HEADERS:
                    candidates_id[h] = ci
            if candidates_name and candidates_id:
                header_row_idx = ri
                name_col = next(iter(candidates_name.values()))
                id_col = next(iter(candidates_id.values()))
                if name_col == id_col:
                    errors.append(
                        "The spreadsheet row labeled as headers maps Name and ID to the same column. "
                        "Use two columns: one for Name and one for ID."
                    )
                    return [], errors
                break

        if header_row_idx is None or name_col is None or id_col is None:
            errors.append(
                "Could not find Name and ID columns. Use headers such as 'Name' and 'ID' (or 'NIM', 'NIP') in the first rows."
            )
            return [], errors

        seen_ids: set[str] = set()
        rows: list[RecipientRow] = []
        data_rows_seen = 0
        for sheet_row_idx, row in enumerate(
            ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
            start=header_row_idx + 1,
        ):
            data_rows_seen += 1
            if data_rows_seen > max_rows:
                errors.append(f"Spreadsheet exceeds maximum of {max_rows} data rows.")
                break
            if not row:
                continue
            name_val = row[name_col] if name_col < len(row) else None
            id_val = row[id_col] if id_col < len(row) else None
            name_str = (str(name_val).strip() if name_val is not None else "") or ""
            id_str = (str(id_val).strip() if id_val is not None else "") or ""
            if not name_str and not id_str:
                continue
            if not name_str or not id_str:
                errors.append(f"Row {sheet_row_idx}: missing name or ID, skipped.")
                continue
            norm_id = id_str.casefold()
            if norm_id in seen_ids:
                errors.append(f"Duplicate ID in spreadsheet: {id_str}")
                continue
            seen_ids.add(norm_id)
            rows.append(RecipientRow(display_name=name_str, id_raw=id_str))
    finally:
        wb.close()

    return rows, errors
