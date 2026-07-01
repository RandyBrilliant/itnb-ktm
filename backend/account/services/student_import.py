"""Parse Excel files for bulk student account creation."""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Any

from django.core.exceptions import ValidationError
from django.core.validators import validate_email

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from account.student_departments import (
    format_student_department_choices,
    resolve_student_department,
)
from account.placeholder_email import synthetic_student_email

ROLE_STUDENT = "STUDENT"
ROLE_ALUMNI = "ALUMNI"

EMAIL_HEADERS = frozenset({"email", "e-mail", "e mail", "surel"})
FULL_NAME_HEADERS = frozenset(
    {
        "full name",
        "full_name",
        "name",
        "nama",
        "nama lengkap",
        "student name",
    }
)
INST_HEADERS = frozenset(
    {
        "institutional id",
        "institutional_id",
        "nim",
        "student id",
        "official id",
        "nomor induk",
        "nim/nip",
    }
)
DEPT_HEADERS = frozenset({"department", "departemen", "jurusan", "program studi", "fakultas"})
PLACE_OF_BIRTH_HEADERS = frozenset(
    {"place of birth", "place_of_birth", "birth place", "tempat lahir", "kota lahir"}
)
DATE_OF_BIRTH_HEADERS = frozenset(
    {"date of birth", "date_of_birth", "birth date", "tanggal lahir", "dob", "tgl lahir"}
)
ROLE_HEADERS = frozenset({"record type", "record_type", "role", "tipe", "tipe rekam", "jenis"})
ALUMNI_YEAR_HEADERS = frozenset(
    {"graduation year", "graduation_year", "alumni year", "alumni_year", "tahun lulus"}
)
PWD_HEADERS = frozenset({"password", "kata sandi", "pass"})

IMPORT_HEADERS = [
    "Email",
    "Full name",
    "Institutional ID",
    "Department",
    "Place of birth",
    "Date of birth",
    "Record type",
    "Graduation year",
    "Password",
]

SAMPLE_STUDENT_ROW = [
    "",
    "Budi Santoso",
    "2021001234",
    "Information Systems",
    "Jakarta",
    "2003-05-15",
    "STUDENT",
    "",
    "",
]

SAMPLE_ALUMNI_ROW = [
    "alumni@example.com",
    "Siti Rahayu",
    "2018005678",
    "Management",
    "Bandung",
    "2000-12-01",
    "ALUMNI",
    2022,
    "",
]


def _norm(cell_value: Any) -> str:
    if cell_value is None:
        return ""
    s = str(cell_value).strip()
    s = re.sub(r"\s+", " ", s)
    return s.lower()


def _cell_str(cell_value: Any) -> str:
    if cell_value is None:
        return ""
    if isinstance(cell_value, float) and cell_value.is_integer():
        return str(int(cell_value))
    return str(cell_value).strip()


def _cell_raw(row: tuple[Any, ...], header_map: dict[str, int], name: str) -> Any:
    if name not in header_map:
        return None
    idx = header_map[name]
    if idx >= len(row):
        return None
    return row[idx]


def _parse_date(cell_value: Any) -> date | None:
    if cell_value is None or cell_value == "":
        return None
    if isinstance(cell_value, datetime):
        return cell_value.date()
    if isinstance(cell_value, date):
        return cell_value

    text = _cell_str(cell_value)
    if not text:
        return None

    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%m/%d/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S.%f",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_role(cell_value: Any) -> str | None:
    text = _cell_str(cell_value).upper()
    if not text:
        return ROLE_STUDENT
    if text in {ROLE_STUDENT, "MAHASISWA"}:
        return ROLE_STUDENT
    if text in {ROLE_ALUMNI, "ALUMNUS", "LULUSAN"}:
        return ROLE_ALUMNI
    return None


def _parse_alumni_year(cell_value: Any) -> int | None:
    if cell_value is None or cell_value == "":
        return None
    if isinstance(cell_value, float) and cell_value.is_integer():
        return int(cell_value)
    text = _cell_str(cell_value)
    if not text:
        return None
    if not text.isdigit():
        return None
    return int(text)


def build_student_import_workbook() -> io.BytesIO:
    """Build a styled .xlsx template with sample rows and an instructions sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(IMPORT_HEADERS)
    ws.append(SAMPLE_STUDENT_ROW)
    ws.append(SAMPLE_ALUMNI_ROW)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="AF0F24")
    for col_idx, header in enumerate(IMPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        width = max(len(header) + 4, 16)
        if header == "Email":
            width = 28
        elif header == "Full name":
            width = 24
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    instructions = wb.create_sheet("Instructions")
    instruction_rows = [
        ["Student record bulk import"],
        [""],
        ["1. Fill in the Students sheet (delete the sample rows before importing)."],
        ["2. Required columns: Full name and Institutional ID (NIM)."],
        [
            "3. Optional columns: Email, Department, Place of birth, Date of birth (YYYY-MM-DD), "
            "Record type (STUDENT or ALUMNI), Graduation year, Password."
        ],
        [
            "   If Email is left empty, a placeholder address is created and students sign in "
            "with their Institutional ID."
        ],
        [f"   Allowed departments: {format_student_department_choices()}."],
        ["4. Leave Password empty to use each row's Institutional ID as the initial password."],
        ["5. Graduation year is only used when Record type is ALUMNI."],
        ["6. Save as .xlsx and upload from Admin > Student records."],
    ]
    for row in instruction_rows:
        instructions.append(row)
    instructions.column_dimensions["A"].width = 100
    instructions["A1"].font = Font(bold=True, size=13)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _resolve_worksheet(wb):
    if "Students" in wb.sheetnames:
        return wb["Students"]
    return wb.active


def parse_student_import_xlsx(
    raw: bytes,
    *,
    max_rows: int = 5_000,
) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Read the Students sheet (or first sheet); row 1 = headers.
    Returns (rows, parse_errors).
    """
    errors: list[str] = []
    rows_out: list[dict[str, Any]] = []

    try:
        wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        return [], [f"Could not read Excel file: {exc}"]

    try:
        ws = _resolve_worksheet(wb)
        header_map: dict[str, int] = {}
        header_row_idx = 1

        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            return [], ["The spreadsheet is empty."]

        for ci, cell in enumerate(first):
            key = _norm(cell)
            if not key:
                continue
            if key in EMAIL_HEADERS:
                header_map["email"] = ci
            elif key in FULL_NAME_HEADERS:
                header_map["full_name"] = ci
            elif key in INST_HEADERS:
                header_map["institutional_id"] = ci
            elif key in DEPT_HEADERS:
                header_map["department"] = ci
            elif key in PLACE_OF_BIRTH_HEADERS:
                header_map["place_of_birth"] = ci
            elif key in DATE_OF_BIRTH_HEADERS:
                header_map["date_of_birth"] = ci
            elif key in ROLE_HEADERS:
                header_map["role"] = ci
            elif key in ALUMNI_YEAR_HEADERS:
                header_map["alumni_year"] = ci
            elif key in PWD_HEADERS:
                header_map["password"] = ci

        if "full_name" not in header_map or "institutional_id" not in header_map:
            return [], [
                "Missing required columns. The first row must include headers for Full name "
                "and Institutional ID (e.g. columns labeled Full name and Institutional ID)."
            ]

        seen_emails: set[str] = set()
        seen_institutional_ids: set[str] = set()

        def col(row: tuple[Any, ...], name: str) -> str:
            if name not in header_map:
                return ""
            idx = header_map[name]
            if idx >= len(row):
                return ""
            return _cell_str(row[idx])

        for ri, row in enumerate(
            ws.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + max_rows, values_only=True),
            start=header_row_idx + 1,
        ):
            if not row:
                continue
            email = col(row, "email")
            full_name = col(row, "full_name")
            inst_raw = col(row, "institutional_id")
            if not full_name and not inst_raw:
                continue
            if not full_name:
                errors.append(f"Row {ri}: missing full name.")
                continue

            inst = inst_raw.strip()
            if not inst:
                errors.append(f"Row {ri}: missing institutional ID.")
                continue

            inst_key = inst.lower()
            if inst_key in seen_institutional_ids:
                errors.append(f"Row {ri}: duplicate institutional ID in file ({inst}).")
                continue
            seen_institutional_ids.add(inst_key)

            email_provided = bool(email.strip())
            if email_provided:
                try:
                    validate_email(email.strip())
                except ValidationError:
                    errors.append(f"Row {ri}: invalid email ({email.strip()}).")
                    continue

            resolved_email = email.strip() if email_provided else synthetic_student_email(inst)
            el = resolved_email.strip().lower()
            if el in seen_emails:
                label = email if email_provided else f"(generated for {inst})"
                errors.append(f"Row {ri}: duplicate email in file ({label}).")
                continue
            seen_emails.add(el)

            role = _parse_role(_cell_raw(row, header_map, "role"))
            if role is None:
                errors.append(
                    f"Row {ri}: invalid record type. Use STUDENT or ALUMNI."
                )
                continue

            dob_raw = _cell_raw(row, header_map, "date_of_birth")
            date_of_birth = None
            if dob_raw not in (None, ""):
                date_of_birth = _parse_date(dob_raw)
                if date_of_birth is None:
                    errors.append(
                        f"Row {ri}: invalid date of birth ({_cell_str(dob_raw)}). Use YYYY-MM-DD."
                    )
                    continue

            alumni_year = None
            alumni_raw = _cell_raw(row, header_map, "alumni_year")
            if alumni_raw not in (None, ""):
                alumni_year = _parse_alumni_year(alumni_raw)
                if alumni_year is None:
                    errors.append(
                        f"Row {ri}: invalid graduation year ({_cell_str(alumni_raw)})."
                    )
                    continue

            if role == ROLE_STUDENT:
                alumni_year = None

            department_raw = col(row, "department") or ""
            department, dept_error = resolve_student_department(department_raw)
            if dept_error:
                errors.append(f"Row {ri}: {dept_error}")
                continue

            rows_out.append(
                {
                    "row": ri,
                    "email": resolved_email,
                    "full_name": full_name.strip(),
                    "institutional_id": inst,
                    "department": department,
                    "place_of_birth": col(row, "place_of_birth") or "",
                    "date_of_birth": date_of_birth,
                    "role": role,
                    "alumni_year": alumni_year,
                    "password": col(row, "password") or None,
                }
            )
    finally:
        wb.close()

    return rows_out, errors
